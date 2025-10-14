import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']

    hrading4=sheet.cell(1,4)
    hrading4.value='newcol'

    for row in range(2,sheet.max_row +1):
        cell=sheet.cell(row,3)
        correctedval=cell.value * 0.9
        correct_cell=sheet.cell(row,4)
        correct_cell.value=correctedval

    values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)

    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)